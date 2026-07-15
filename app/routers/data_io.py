"""Import / Export — Excel (.xlsx) with 3 sheets.

Export: single .xlsx file with sheets 车辆 / 加油记录 / 保养记录.
Import: incremental upsert, vehicle association by name, data validation.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlmodel import Session

from app.db import get_session
from app.models.fuel_record import FuelRecord
from app.models.maintenance import MaintenanceRecord
from app.models.vehicle import Vehicle
from app.security import CurrentUser, verify_token
from app.services.helpers import gen_id

router = APIRouter(prefix="/api/v1/data", tags=["data"])

# ── Column definitions (Chinese headers) ──────────────────────────────────

VEHICLE_HEADERS = ["车辆名称", "车牌号", "车型"]
RECORD_HEADERS = [
    "车辆名称", "日期", "里程(公里)", "加油量(升)", "单价(元/升)",
    "是否加满", "油品", "加油站", "备注", "油量灯亮",
    "上次加油没记录", "机显金额", "实付金额",
]
MAINT_HEADERS = [
    "车辆名称", "日期", "里程(公里)", "保养类型", "自定义名称",
    "项目", "费用(元)", "备注", "提醒方式", "下次日期", "下次里程(公里)",
]

TRIGGER_MAP = {"日期": "date", "里程": "odo", "都提醒": "either", "不提醒": "none"}
TRIGGER_REVERSE = {v: k for k, v in TRIGGER_MAP.items()}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _safe_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _style_header(ws, headers: list[str]):
    """Write header row with styling."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                # Chinese chars count as ~2 width units.
                clen = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, clen)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ── Export ─────────────────────────────────────────────────────────────────

@router.get("/export")
def export_xlsx(
    current: CurrentUser = Depends(verify_token),
    session: Session = Depends(get_session),
):
    vehicles = list(
        session.execute(select(Vehicle).where(Vehicle.user_id == current.id))
        .scalars().all()
    )
    vname_map = {v.id: v.name for v in vehicles}
    vids = {v.id for v in vehicles}

    records = list(
        session.execute(select(FuelRecord).where(FuelRecord.vehicle_id.in_(vids or {""})))
        .scalars().all()
    )
    maints = list(
        session.execute(select(MaintenanceRecord).where(MaintenanceRecord.vehicle_id.in_(vids or {""})))
        .scalars().all()
    )

    wb = Workbook()

    # Sheet 1: 车辆
    ws_v = wb.active
    ws_v.title = "车辆"
    _style_header(ws_v, VEHICLE_HEADERS)
    for v in vehicles:
        ws_v.append([v.name, v.plate, v.model])
    _auto_width(ws_v)

    # Sheet 2: 加油记录
    ws_r = wb.create_sheet("加油记录")
    _style_header(ws_r, RECORD_HEADERS)
    for r in records:
        ws_r.append([
            vname_map.get(r.vehicle_id, ""),
            str(r.record_date)[:10],
            float(r.odometer),
            float(r.liters),
            float(r.price),
            "加满" if r.full_tank == "yes" else "未加满",
            r.fuel_type,
            r.station,
            r.note,
            "是" if r.light else "否",
            "是" if r.skipped_previous else "否",
            float(r.pump_amount) if r.pump_amount else "",
            float(r.paid_amount) if r.paid_amount else "",
        ])
    _auto_width(ws_r)

    # Sheet 3: 保养记录
    ws_m = wb.create_sheet("保养记录")
    _style_header(ws_m, MAINT_HEADERS)
    for m in maints:
        ws_m.append([
            vname_map.get(m.vehicle_id, ""),
            str(m.record_date)[:10],
            float(m.odometer),
            m.maint_type,
            m.custom_name,
            m.item,
            float(m.cost),
            m.note,
            TRIGGER_REVERSE.get(m.trigger, "都提醒"),
            str(m.next_date)[:10] if m.next_date else "",
            float(m.next_odo) if m.next_odo else "",
        ])
    _auto_width(ws_m)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"省油的灯_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Import ─────────────────────────────────────────────────────────────────

def _read_sheet(wb, sheet_name: str, headers: list[str]) -> list[dict]:
    """Read a sheet into a list of dicts keyed by Chinese header names."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    result = []
    for row in rows:
        d = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            d[h] = str(val).strip() if val is not None else ""
        result.append(d)
    return result


def _validate_import(
    v_rows: list[dict],
    r_rows: list[dict],
    m_rows: list[dict],
) -> list[str]:
    errors: list[str] = []
    vnames: set[str] = set()

    for i, v in enumerate(v_rows, 2):
        name = v.get("车辆名称", "")
        if not name:
            errors.append(f"车辆 第{i}行：车辆名称不能为空")
        elif name in vnames:
            errors.append(f"车辆 第{i}行：车辆名称「{name}」重复")
        vnames.add(name)

    for i, r in enumerate(r_rows, 2):
        vname = r.get("车辆名称", "")
        if not vname:
            errors.append(f"加油记录 第{i}行：车辆名称不能为空")
        elif vname not in vnames:
            errors.append(f"加油记录 第{i}行：车辆「{vname}」不存在于车辆表中")

        d = r.get("日期", "")
        if not d:
            errors.append(f"加油记录 第{i}行：日期不能为空")
        else:
            try:
                date.fromisoformat(d[:10])
            except ValueError:
                errors.append(f"加油记录 第{i}行：日期格式错误「{d}」")

        odo = _safe_decimal(r.get("里程(公里)"))
        if odo is None or odo < 0:
            errors.append(f"加油记录 第{i}行：里程必须为非负数")

        liters = _safe_decimal(r.get("加油量(升)"))
        if liters is None or liters <= 0:
            errors.append(f"加油记录 第{i}行：加油量必须为正数")

        price = _safe_decimal(r.get("单价(元/升)"))
        if price is None or price <= 0:
            errors.append(f"加油记录 第{i}行：单价必须为正数")

        ft = r.get("是否加满", "")
        if ft and ft not in ("加满", "未加满"):
            errors.append(f"加油记录 第{i}行：是否加满 应为「加满」或「未加满」")

    for i, m in enumerate(m_rows, 2):
        vname = m.get("车辆名称", "")
        if not vname:
            errors.append(f"保养记录 第{i}行：车辆名称不能为空")
        elif vname not in vnames:
            errors.append(f"保养记录 第{i}行：车辆「{vname}」不存在于车辆表中")

    return errors


@router.post("/import")
async def import_xlsx(
    file: UploadFile,
    current: CurrentUser = Depends(verify_token),
    session: Session = Depends(get_session),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 Excel 文件（.xlsx）")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "无法读取 Excel 文件，请确认格式正确")

    v_rows = _read_sheet(wb, "车辆", VEHICLE_HEADERS)
    r_rows = _read_sheet(wb, "加油记录", RECORD_HEADERS)
    m_rows = _read_sheet(wb, "保养记录", MAINT_HEADERS)
    wb.close()

    if not v_rows and not r_rows and not m_rows:
        raise HTTPException(400, "Excel 中未找到有效数据（需要 车辆/加油记录/保养记录 工作表）")

    errors = _validate_import(v_rows, r_rows, m_rows)
    if errors:
        raise HTTPException(422, detail={"message": f"数据校验失败（{len(errors)} 项）", "errors": errors[:30]})

    # Build vehicle name → id mapping.
    existing_vehicles = {
        v.name: v for v in
        session.execute(select(Vehicle).where(Vehicle.user_id == current.id))
        .scalars().all()
    }
    vname_to_id: dict[str, str] = {}
    counts = {"vehicles": 0, "records": 0, "maint": 0}

    # Upsert vehicles.
    for v in v_rows:
        name = v["车辆名称"]
        if name in existing_vehicles:
            veh = existing_vehicles[name]
            veh.plate = v.get("车牌号", veh.plate)
            veh.model = v.get("车型", veh.model)
            vname_to_id[name] = veh.id
        else:
            vid = gen_id("v")
            session.add(Vehicle(
                id=vid, user_id=current.id,
                name=name, plate=v.get("车牌号", ""), model=v.get("车型", ""),
            ))
            vname_to_id[name] = vid
        counts["vehicles"] += 1
    session.flush()

    # Upsert fuel records (dedup: vehicle + date + odometer).
    for r in r_rows:
        vid = vname_to_id.get(r["车辆名称"], "")
        if not vid:
            continue
        d = r["日期"][:10]
        odo = _safe_decimal(r.get("里程(公里)")) or Decimal("0")
        liters = _safe_decimal(r.get("加油量(升)")) or Decimal("0")
        price = _safe_decimal(r.get("单价(元/升)")) or Decimal("0")
        total = round(liters * price, 3)
        pump = _safe_decimal(r.get("机显金额")) or total
        paid = _safe_decimal(r.get("实付金额")) or pump
        ft = "yes" if r.get("是否加满") == "加满" else "no"

        existing = session.execute(
            select(FuelRecord).where(
                FuelRecord.vehicle_id == vid,
                FuelRecord.record_date == d,
                FuelRecord.odometer == odo,
            )
        ).scalars().first()

        if existing:
            existing.liters = liters
            existing.price = price
            existing.total_cost = total
            existing.pump_amount = pump
            existing.paid_amount = paid
            existing.full_tank = ft
            existing.fuel_type = r.get("油品", "92")
            existing.station = r.get("加油站", "")
            existing.note = r.get("备注", "")
            existing.light = r.get("油量灯亮") == "是"
            existing.skipped_previous = r.get("上次加油没记录") == "是"
        else:
            session.add(FuelRecord(
                id=gen_id("r"), vehicle_id=vid,
                record_date=d, odometer=odo,
                liters=liters, price=price, total_cost=total,
                pump_amount=pump, paid_amount=paid,
                full_tank=ft,
                fuel_type=r.get("油品", "92"),
                station=r.get("加油站", ""),
                note=r.get("备注", ""),
                light=r.get("油量灯亮") == "是",
                skipped_previous=r.get("上次加油没记录") == "是",
            ))
        counts["records"] += 1

    # Upsert maintenance records (dedup: vehicle + date + item).
    for m in m_rows:
        vid = vname_to_id.get(m["车辆名称"], "")
        if not vid:
            continue
        d = m.get("日期", "")[:10]
        item = m.get("项目", "")

        existing = session.execute(
            select(MaintenanceRecord).where(
                MaintenanceRecord.vehicle_id == vid,
                MaintenanceRecord.record_date == d,
                MaintenanceRecord.item == item,
            )
        ).scalars().first()

        trigger = TRIGGER_MAP.get(m.get("提醒方式", ""), "either")
        nd = m.get("下次日期", "")
        next_date = nd[:10] if nd else None
        next_odo = _safe_decimal(m.get("下次里程(公里)"))

        if existing:
            existing.odometer = _safe_decimal(m.get("里程(公里)")) or Decimal("0")
            existing.maint_type = m.get("保养类型", "")
            existing.custom_name = m.get("自定义名称", "")
            existing.cost = _safe_decimal(m.get("费用(元)")) or Decimal("0")
            existing.note = m.get("备注", "")
            existing.trigger = trigger
            existing.next_date = next_date
            existing.next_odo = next_odo
        else:
            session.add(MaintenanceRecord(
                id=gen_id("m"), vehicle_id=vid,
                record_date=d,
                odometer=_safe_decimal(m.get("里程(公里)")) or Decimal("0"),
                maint_type=m.get("保养类型", ""),
                custom_name=m.get("自定义名称", ""),
                item=item,
                cost=_safe_decimal(m.get("费用(元)")) or Decimal("0"),
                note=m.get("备注", ""),
                trigger=trigger,
                next_date=next_date,
                next_odo=next_odo,
            ))
        counts["maint"] += 1

    return counts
